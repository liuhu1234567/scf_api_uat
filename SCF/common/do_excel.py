from openpyxl import load_workbook
from config.all_path import get_file_path


class DoExcel(object):
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.file_path = get_file_path(file_name)
        self.wb = load_workbook(self.file_path)
        self.ws = self.wb[sheet_name]

    def insert(self, all_value, row_start):
        for index, value in enumerate(all_value):
            self.ws.cell(row_start, index + 1, value)
            if self.ws.cell(row_start, 8).value is not None:
                trg = 'H' + str(row_start)
                self.ws[trg].number_format = 'yyyy/m/d'


    def save(self):
        self.file_name_l = self.file_name.split('.')
        self.new_file_name = f'{self.file_name_l[0]}_new.{self.file_name_l[1]}'
        self.wb.save(get_file_path(self.new_file_name))
        print(f'excel保存成功：{self.new_file_name}')
        return self.new_file_name


if __name__ == '__main__':
    excel = DoExcel('批量导入客户.xlsx', '批量邀请客户建档模板')
    l = [1, 1, 1, 1, 1, 1, 1, 1]
    excel.insert(l, 3)
    excel.save()
